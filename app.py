import os, base64, pickle, io, img2pdf, re, json
from flask import Flask, request, render_template, jsonify
from google.cloud import vision
from google import genai
from google.genai import types
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import load_workbook


app = Flask(__name__)


# --- Railwayの環境変数からGCP認証ファイルを生成するロジック ---
# これにより、Google Cloud SDK が自動的にこのファイルを見にいきます
if os.environ.get("GOOGLE_APPLICATION_CREDENTIALS_JSON"):
    key_path = os.path.join(os.getcwd(), "gcp-key.json")
    with open(key_path, "w") as f:
        f.write(os.environ.get("GOOGLE_APPLICATION_CREDENTIALS_JSON"))
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = key_path
    print(f"GCP Key file created at: {key_path}")


# 新しい SDK のクライアント初期化
client_gemini = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))

# ===========================
# 設定値
# ===========================

TOKEN_PICKLE_B64 = "gASV8QMAAAAAAACMGWdvb2dsZS5vYXV0aDIuY3JlZGVudGlhbHOUjAtDcmVkZW50aWFsc5STlCmBlH2UKIwFdG9rZW6UjP15YTI5LmEwQVVNV2dfTGFYeTQ5dC1pb082Y1JrRVMxcGJhRUkxUE5HVnlkMlBNZnA2MGQtMUtOdWdIV0VwejFiS0NYU0JvVFY3aEtWT19NektTTUdLbV9lQ2dPd1J5UG9IT2RiTk5WQV9lTmF5cjNUMlhUaDd1Nmx0Z0FNTkNPcDYyV2hOdlA4bHNzbnlPbEdrc0RKNFZCRWowZzE4UVBVY0pCTUNFZDg0UTZvWVFKZVZzaTlhb2J6VUM0bnAtaFQyZ3RjVVRvc3pNWG10c2FDZ1lLQVZZU0FROFNGUUhHWDJNaUctMWZLQjlreXk5cDlOTk1CdW4wQVEwMjA2lIwGZXhwaXJ5lIwIZGF0ZXRpbWWUjAhkYXRldGltZZSTlEMKB+oBEA4TOAAAAJSFlFKUjBFfcXVvdGFfcHJvamVjdF9pZJROjA9fdHJ1c3RfYm91bmRhcnmUTowQX3VuaXZlcnNlX2RvbWFpbpSMDmdvb2dsZWFwaXMuY29tlIwZX3VzZV9ub25fYmxvY2tpbmdfcmVmcmVzaJSJjAdfc2NvcGVzlF2UjCVodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL2RyaXZllGGMD19kZWZhdWx0X3Njb3Blc5ROjA5fcmVmcmVzaF90b2tlbpSMZzEvLzBncEVOY2NoYmZCNXRDZ1lJQVJBQUdCQVNOd0YtTDlJcmtBYS1EajRBWm1pRVQwMGYyNVN3bE5VNU55MFo3X3ZLUEFXdi1oVnd0aXNNRXNBUDZDWGR0cWdLNnBseGNmenBKOHeUjAlfaWRfdG9rZW6UTowPX2dyYW50ZWRfc2NvcGVzlF2UjCVodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL2RyaXZllGGMCl90b2tlbl91cmmUjCNodHRwczovL29hdXRoMi5nb29nbGVhcGlzLmNvbS90b2tlbpSMCl9jbGllbnRfaWSUjEg3Mzg1NzkzMzc0NTMtaTRiM2toYXA2ZjEwcmlybzhqOGM3ZmZqZnJoNGUzYzAuYXBwcy5nb29nbGV1c2VyY29udGVudC5jb22UjA5fY2xpZW50X3NlY3JldJSMI0dPQ1NQWC1jcnljX0JVWmM5VFlqMWNtRTNzajZVcmZHczZ6lIwLX3JhcHRfdG9rZW6UTowWX2VuYWJsZV9yZWF1dGhfcmVmcmVzaJSJjAhfYWNjb3VudJSMAJSMD19jcmVkX2ZpbGVfcGF0aJROdWIu"
EXCEL_FILE_ID = "1rf3DTxGpTNM0VZxcBkMjV2AyhE0oDiJlgv-_V_G3pbk"      # Excel ファイルID
RECEIPTS_FOLDER_ID = "1UaC4E-5O408ozxKx_VlFoYWilFWTbf-f"           # Drive フォルダID
SCOPES = ['https://www.googleapis.com/auth/drive']


# ===========================
# OCR解析用エンドポイント
# ===========================
@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        if "receipt" not in request.files:
            return jsonify({"error": "No file"}), 400
        
        file = request.files["receipt"]
        content = file.read()

        # 1. Vision APIで文字抽出
        client_vision = vision.ImageAnnotatorClient()
        image = vision.Image(content=content)
        response_vision = client_vision.text_detection(image=image)
        
        if not response_vision.text_annotations:
            return jsonify({"pay_date": "", "payee": "", "amount": ""})
            
        full_text = response_vision.text_annotations[0].description

        # 2. 新しい SDK を使用した Gemini 解析
        prompt = f"""
        以下の領収書の解析テキストから、[支払日(YYYY-MM-DD形式), 支払先名称, 合計金額(数値のみ)]を抽出し、
        必ず以下のJSON形式のみで回答してください。和暦は西暦に変換してください。
        不明な項目は空文字にしてください。
        {{
          "pay_date": "YYYY-MM-DD",
          "payee": "店名",
          "amount": "1234"
        }}
        テキスト:
        {full_text}
        """
        
        # 新しいメソッド呼び出し形式
        response_gemini = client_gemini.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt
        )
        
        # JSON部分を抽出してパース
        json_match = re.search(r'\{.*\}', response_gemini.text, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            return jsonify(result)
        else:
            return jsonify({"pay_date": "", "payee": "", "amount": ""})

    except Exception as e:
        print(f"AI OCR Error: {e}")
        return jsonify({"error": str(e)}), 500
         

# ===========================
# Google Drive サービス作成
# ===========================
def get_drive_service():
    if TOKEN_PICKLE_B64:
        token_bytes = base64.b64decode(TOKEN_PICKLE_B64)
        creds = pickle.load(io.BytesIO(token_bytes))
        service = build('drive', 'v3', credentials=creds)
        return service
    else:
        raise Exception("Google API credentials are missing")

# ===========================
# Excelに追記
# ===========================
def update_excel(service, filename, pay_date, payee, amount):
    # 1. スプレッドシートをExcel形式でエクスポートして取得
    request_dl = service.files().export_media(
        fileId=EXCEL_FILE_ID,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request_dl)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.seek(0)

    # 2. openpyxlで読み込み
    wb = load_workbook(fh)
    ws = wb.active

    # 3. 手動編集による「空行」対策：データが入っている本当の最終行を探す
    real_last_row = 0
    for row in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row]):
            real_last_row = row
            break
    
    # 本当の最終行の次に追加
    new_row = real_last_row + 1
    ws.cell(row=new_row, column=1, value=pay_date)
    ws.cell(row=new_row, column=2, value=payee)
    ws.cell(row=new_row, column=3, value=amount)
    ws.cell(row=new_row, column=4, value=filename)

    # 4. メモリ上のバイナリに保存
    out_fh = io.BytesIO()
    wb.save(out_fh)
    out_fh.seek(0)

    # 5. 【重要】Googleドライブ上の既存ファイルを更新
    # スプレッドシート形式を維持したまま上書きする
    media = MediaIoBaseUpload(
        out_fh, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True
    )
    # updateメソッドで、既存のファイルIDを指定して中身を入れ替える
    service.files().update(
        fileId=EXCEL_FILE_ID,
        media_body=media,
        supportsAllDrives=True
    ).execute()
    
# ===========================
# ファイルアップロード（PDF変換版）
# ===========================
def upload_file_to_drive(service, file, filename):
    # ファイル名を .jpg から .pdf に変更
    pdf_filename = filename.replace(".jpg", ".pdf")
    
    # 画像をPDFに変換
    pdf_bytes = img2pdf.convert(file.stream)
    pdf_stream = io.BytesIO(pdf_bytes)

    file_metadata = {
        "name": pdf_filename, 
        "parents": [RECEIPTS_FOLDER_ID]
    }
    
    # MIMEタイプを application/pdf に指定
    media = MediaIoBaseUpload(
        pdf_stream, 
        mimetype="application/pdf"
    )
    
    # 【変更】fields='name, webViewLink' を指定して、URLを取得する
    uploaded_file = service.files().create(
        body=file_metadata, 
        media_body=media, 
        fields='name, webViewLink',
        supportsAllDrives=True
    ).execute()
    
    return uploaded_file.get('name'), uploaded_file.get('webViewLink')

# ===========================
# ルート
# ===========================

@app.route("/", methods=["GET", "POST"])
def index():
    message = ""
    if request.method == "POST":
        if "receipt" not in request.files:
            message = "画像が送信されていません。"
            return render_template("index.html", message=message)

        file = request.files["receipt"]
        if file.filename == "":
            message = "ファイル名がありません。"
            return render_template("index.html", message=message)

        pay_date = request.form.get("pay_date")
        payee = request.form.get("payee")
        amount = request.form.get("amount")
        
        # ベースとなるファイル名（拡張子なし）
        base_filename = f"{payee}_{pay_date}_{amount}"

        # ファイル名作成
        filename = f"{payee} {pay_date} {amount}.jpg"

        try:
            service = get_drive_service()
            
          # 【変更】URLも受け取る
            final_filename, file_url = upload_file_to_drive(service, file, base_filename)

            
            # ExcelにはPDFのファイル名を記録
            update_excel(service, final_filename, pay_date, payee, amount)

            # スプレッドシートのURLを生成
            spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{EXCEL_FILE_ID}/edit"
            
            # 【変更】成功時にURLを含むJSONを返す
            return jsonify({
                "status": "success",
                "file_url": file_url,
                "spreadsheet_url": spreadsheet_url
            }), 200
            
            
        except Exception as e:
            message = f"エラー: {e}"
            print(f"Error detail: {e}") # サーバーログに詳細を出力
            return "Internal Server Error", 500 # JavaScript側に失敗を伝える

    return render_template("index.html", message=message)

# ===========================
if __name__ == "__main__":
    # Railway などの環境でポート番号を正しく取得する設定
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
    
